import logging
import re
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from dateutil import parser as dateparser

from database import get_session, Processo
from datajud_client import atualizar_processo

logger = logging.getLogger(__name__)


def _extrair_cnj(texto: str) -> str:
    padrao = r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}"
    match = re.search(padrao, texto)
    if match:
        return match.group(0)
    padrao_flex = r"\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{3,6}"
    match = re.search(padrao_flex, texto)
    if match:
        return match.group(0)
    numbers = re.findall(r"\d+", texto)
    cnj = "".join(numbers)
    if len(cnj) >= 20:
        return cnj[:25]
    return ""


MAPEAMENTO = {
    "processo": ["processo", "número", "numero", "nº", "n°", "n.º", "cnj", "num_cnj", "num_proc", "númeroprocesso", "numeroprocesso", "protocolo"],
    "tribunal": ["tribunal", "orgao", "órgão", "vara", "comarca", "tribunalorigem"],
    "classe": ["classe", "classe_principal", "natureza", "classejudicial"],
    "assunto": ["assunto", "assunto_principal", "assuntos"],
    "autora": ["autora", "autor", "requerente", "parte_ativa", "parte_autora", "requerentes"],
    "re": ["reu", "ré", "requerido", "parte_passiva", "parte_re", "requeridos"],
    "oab": ["oab", "advogado", "adv", "oab_advogado", "advogados"],
    "data_ajuizamento": ["data_ajuizamento", "data_distribuicao", "data_autuacao", "data", "dataajuizamento"],
}


def _mapear_colunas(ws) -> dict:
    colunas = {}
    for cell in ws[1]:
        colunas[cell.value] = cell.column

    resultado = {}
    for campo, aliases in MAPEAMENTO.items():
        col_idx = None
        for alias in aliases:
            for col_nome, col_idx_candidate in colunas.items():
                if col_nome and alias in str(col_nome).lower().strip():
                    col_idx = col_idx_candidate
                    break
            if col_idx:
                break
        resultado[campo] = col_idx

    return resultado, list(colunas.keys())


def _extrair_cnj_da_linha(valores: list) -> str:
    texto = " ".join(str(v) for v in valores if v)
    return _extrair_cnj(texto)


def _importar_xlsx_astrea(ws, session) -> dict:
    importados = 0
    importados_cnjs = []
    erros = 0
    ja_existem = 0
    sem_cnj = 0
    erros_amostra = []
    colunas_planilha = []
    amostra_linha = {}

    for i, cell in enumerate(ws[1]):
        colunas_planilha.append(str(cell.value) if cell.value is not None else f"Col{i+1}")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False)):
        try:
            valores = [cell.value for cell in row]

            if row_idx == 0:
                for ci, v in enumerate(valores):
                    amostra_linha[f"col_{ci+1}"] = str(v)[:80] if v is not None else "(vazio)"

            cnj = _extrair_cnj_da_linha(valores)
            if not cnj:
                sem_cnj += 1
                continue

            existe = session.query(Processo).filter_by(numero_cnj=cnj).first()
            if existe:
                ja_existem += 1
                continue

            tribunal = ""
            classe = ""
            assunto = ""
            autora = ""
            reu = ""
            oab = ""
            data_ajuizamento = None

            for i, val in enumerate(valores):
                if not val:
                    continue
                val_lower = str(val).lower().strip()

                if i == 0 and val_lower == "processo":
                    caption = str(valores[i + 1]) if i + 1 < len(valores) and valores[i + 1] else ""
                    if " x " in caption:
                        parts = caption.split(" x ", 1)
                        autora = parts[0].strip()
                        reu = parts[1].strip()
                    continue

                if val_lower in ("autor", "autores", "requerente") and i + 1 < len(valores) and valores[i + 1]:
                    if not autora:
                        autora = str(valores[i + 1])
                    continue

                if val_lower in ("réu", "reu", "requerido", "ré", "re") and i + 1 < len(valores) and valores[i + 1]:
                    if not reu:
                        reu = str(valores[i + 1])
                    continue

                if val_lower in ("classe", "natureza", "classe_principal"):
                    if i + 1 < len(valores) and valores[i + 1]:
                        classe = str(valores[i + 1])
                    continue

                if val_lower in ("assunto", "assuntos"):
                    if i + 1 < len(valores) and valores[i + 1]:
                        assunto = str(valores[i + 1])
                    continue

                if val_lower == "tribunal":
                    if i + 1 < len(valores) and valores[i + 1]:
                        tribunal = str(valores[i + 1])
                    continue

                if val_lower in ("data", "data_ajuizamento", "data distribuição", "data autuação", "data da distribuição"):
                    if i + 1 < len(valores) and valores[i + 1]:
                        vd = valores[i + 1]
                        if isinstance(vd, datetime):
                            data_ajuizamento = vd
                        elif isinstance(vd, str):
                            try:
                                data_ajuizamento = dateparser.parse(vd, dayfirst=True)
                            except (ValueError, TypeError):
                                pass
                    continue

                if val_lower in ("oab", "advogado", "advogados"):
                    if i + 1 < len(valores) and valores[i + 1]:
                        oab = str(valores[i + 1])
                    continue

                match_oab = re.search(r"[A-Z]{2}\d{6}", str(val))
                if match_oab and not oab:
                    oab = match_oab.group(0)

            processo = Processo(
                numero_cnj=cnj,
                tribunal=tribunal or "",
                classe=classe or "",
                assunto=assunto or "",
                parte_autora=autora or "",
                parte_re=reu or "",
                advogado_oab=oab or "",
                data_ajuizamento=data_ajuizamento,
            )
            session.add(processo)
            importados += 1
            importados_cnjs.append(cnj)

            if importados % 50 == 0:
                session.commit()
        except Exception as e:
            logger.error(f"Erro ao importar linha Astrea: {e}")
            erros += 1
            if len(erros_amostra) < 3:
                erros_amostra.append(str(e))

    session.commit()
    return {
        "importados": importados,
        "importados_cnjs": importados_cnjs,
        "erros": erros,
        "ja_existem": ja_existem,
        "sem_cnj": sem_cnj,
        "erros_amostra": erros_amostra,
        "colunas_planilha": colunas_planilha,
        "amostra_linha": amostra_linha,
        "formato": "astrea",
        "mensagem": _mensagem_resultado(importados, erros, ja_existem, sem_cnj, "astrea"),
    }


def _mensagem_resultado(importados: int, erros: int, ja_existem: int, sem_cnj: int, formato: str) -> str:
    if importados > 0:
        return ""
    if ja_existem > 0 and erros == 0 and sem_cnj == 0:
        return f"Todos os {ja_existem} processos já estavam cadastrados"
    if ja_existem == 0 and erros == 0 and sem_cnj == 0:
        return "Nenhum processo encontrado nas linhas da planilha"
    return ""


def _importar_xlsx_normal(ws, session) -> dict:
    colunas_idx, colunas_planilha = _mapear_colunas(ws)
    col_processo = colunas_idx["processo"]

    colunas_mapeadas = {}
    for campo in MAPEAMENTO:
        colunas_mapeadas[campo] = colunas_idx.get(campo) is not None

    amostra_linha = {}
    for row_sample in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for i, val in enumerate(row_sample):
            amostra_linha[f"col_{i+1}"] = str(val)[:80] if val is not None else "(vazio)"
        break

    if col_processo is None:
        return {
            "status": "erro",
            "mensagem": "Coluna 'Processo' (CNJ) não encontrada na planilha",
            "colunas_planilha": colunas_planilha,
            "colunas_mapeadas": colunas_mapeadas,
            "amostra_linha": amostra_linha,
            "dica": "Verifique se a planilha tem uma coluna com nome: processo, número, CNJ, num_cnj ou nº",
        }

    importados = 0
    importados_cnjs = []
    erros = 0
    ja_existem = 0
    sem_cnj = 0
    erros_amostra = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        try:
            valor = row[col_processo - 1].value
            if not valor:
                continue

            cnj = _extrair_cnj(str(valor))
            if not cnj:
                sem_cnj += 1
                continue

            existe = session.query(Processo).filter_by(numero_cnj=cnj).first()
            if existe:
                ja_existem += 1
                continue

            tribunal = row[colunas_idx["tribunal"] - 1].value if colunas_idx["tribunal"] else ""
            classe = row[colunas_idx["classe"] - 1].value if colunas_idx["classe"] else ""
            assunto = row[colunas_idx["assunto"] - 1].value if colunas_idx["assunto"] else ""
            autora = row[colunas_idx["autora"] - 1].value if colunas_idx["autora"] else ""
            reu = row[colunas_idx["re"] - 1].value if colunas_idx["re"] else ""
            oab = row[colunas_idx["oab"] - 1].value if colunas_idx["oab"] else ""

            data_ajuizamento = None
            if colunas_idx["data_ajuizamento"]:
                val_data = row[colunas_idx["data_ajuizamento"] - 1].value
                if val_data:
                    if isinstance(val_data, datetime):
                        data_ajuizamento = val_data
                    elif isinstance(val_data, str):
                        try:
                            data_ajuizamento = dateparser.parse(val_data, dayfirst=True)
                        except (ValueError, TypeError):
                            pass

            processo = Processo(
                numero_cnj=cnj,
                tribunal=tribunal or "",
                classe=classe or "",
                assunto=assunto or "",
                parte_autora=autora or "",
                parte_re=reu or "",
                advogado_oab=oab or "",
                data_ajuizamento=data_ajuizamento,
            )
            session.add(processo)
            importados += 1
            importados_cnjs.append(cnj)

            if importados % 50 == 0:
                session.commit()
        except Exception as e:
            logger.error(f"Erro ao importar linha: {e}")
            erros += 1
            if len(erros_amostra) < 3:
                erros_amostra.append(str(e))

    session.commit()
    resultado: dict[str, object] = {
        "status": "ok",
        "importados": importados,
        "importados_cnjs": importados_cnjs,
        "erros": erros,
        "ja_existem": ja_existem,
        "sem_cnj": sem_cnj,
        "colunas_planilha": colunas_planilha,
        "colunas_mapeadas": colunas_mapeadas,
        "amostra_linha": amostra_linha,
    }
    if erros_amostra:
        resultado["erros_amostra"] = erros_amostra
    msg = _mensagem_resultado(importados, erros, ja_existem, sem_cnj, "tabular")
    if msg:
        resultado["mensagem"] = msg
    return resultado


def importar_xlsx(caminho: str) -> dict[str, object]:
    path = Path(caminho)
    if not path.exists():
        return {"status": "erro", "mensagem": f"Arquivo não encontrado: {caminho}"}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    is_astrea = False
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] and str(row[0]).strip().lower() == "processo":
            is_astrea = True
        break

    wb.close()

    session = get_session()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    try:
        if is_astrea:
            resultado = _importar_xlsx_astrea(ws, session)
        else:
            resultado = _importar_xlsx_normal(ws, session)
        resultado.setdefault("status", "ok")
        resultado.setdefault("formato", "tabular")
        resultado.setdefault("colunas_mapeadas", {})

        cnjs_datajud = resultado.get("importados_cnjs", [])
        if cnjs_datajud:
            resultado["datajud_agendados"] = len(cnjs_datajud)

            def _rodar_datajud_em_background():
                for cnj in cnjs_datajud:
                    try:
                        atualizar_processo(cnj)
                    except Exception:
                        pass

            t = threading.Thread(target=_rodar_datajud_em_background, daemon=True)
            t.start()

        return resultado
    finally:
        wb.close()
        session.close()
